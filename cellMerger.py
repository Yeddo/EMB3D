import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import argparse

def merge_adjacent_cells(ws, start_col, end_col):
    """
    Merge adjacent like cells in the given column range.
    Center the text vertically within the merged cells.
    """
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        prev_value = None
        start_row = 2  # Assuming first row is header

        for row in range(2, ws.max_row + 1):  # Start from second row
            current_value = ws[f"{col_letter}{row}"].value

            if current_value == prev_value:
                continue  # Continue merging process

            if prev_value is not None and start_row < row - 1:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=row-1, end_column=col)
                ws[f"{col_letter}{start_row}"].alignment = Alignment(vertical="center", horizontal="center")

            prev_value = current_value
            start_row = row

        # Merge the last sequence if necessary
        if prev_value is not None and start_row < ws.max_row:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=ws.max_row, end_column=col)
            ws[f"{col_letter}{start_row}"].alignment = Alignment(vertical="center", horizontal="center")


def process_csv(file_path):
    # Read CSV
    df = pd.read_csv(file_path)

    # Convert DataFrame to Excel
    excel_path = "output.xlsx"
    df.to_excel(excel_path, index=False, engine='openpyxl')

    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Merge cells in the first four columns
    merge_adjacent_cells(ws, start_col=1, end_col=4)

    # Save the modified Excel file
    wb.save(excel_path)
    print(f"Processed file saved as {excel_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Merge adjacent like cells in the first 4 columns of a CSV and center vertically.")
    parser.add_argument("-f", "--file", required=True, help="Path to the input CSV file")
    
    args = parser.parse_args()
    process_csv(args.file)