"""
MITRE EMB3D JSON Parser with Optional Excel Formatting
Author: Jason Bisnette
Date: March 2025

Description:
    - Used to rebuild mapping spreadsheet when MITRE releases updates to EMB3D
    - Extracts Properties (PID) with descriptions
    - Extracts Threats (TID) with descriptions
    - Extracts Mitigations (MID) with descriptions
    - Maps PIDs --> TIDs and TIDs --> MIDs
    - Optional Excel formatting with `-Format` flag

Usage:
    python3 emb3d_jsonParser.py [-Format]

Dependencies:
    - pandas
    - requests
    - openpyxl
"""

# Import libraries
import os
import sys
import json
import subprocess
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# List of required Python packages
REQUIRED_PACKAGES = ["pandas", "requests", "openpyxl"]

# Function to check required packages/dependencies
def check_and_install_dependencies():
    """ Check for missing Python packages and install them if necessary. """
    missing_packages = []

    # Check if each required package is installed
    for package in REQUIRED_PACKAGES:
        try:
            __import__(package)  # Attempt to import the package
        except ImportError:
            missing_packages.append(package)  # If missing, add to the list

    # Install missing packages using pip
    if missing_packages:
        print(f"Installing missing dependencies: {', '.join(missing_packages)}")
        subprocess.run([sys.executable, "-m", "pip", "install"] + missing_packages, check=True)

# Ensure all dependencies are installed before proceeding
check_and_install_dependencies()

# URLs pointing to MITRE EMB3D JSON data files
THREATS_JSON_URL = "https://raw.githubusercontent.com/mitre/emb3d/main/_data/threats.json"
MITIGATIONS_JSON_URL = "https://raw.githubusercontent.com/mitre/emb3d/main/_data/mitigations_threat_mappings.json"
PROPERTIES_JSON_URL = "https://raw.githubusercontent.com/mitre/emb3d/main/_data/properties_threat_mappings.json"

# Function to fetch .json from MITRE EMB3D github
def fetch_json_from_github(url):
    # Fetch JSON data from a given GitHub URL.
    try:
        response = requests.get(url, timeout=10)  # Send GET request with a timeout
        response.raise_for_status()  # Raise exception if request fails
        return response.json()  # Return the parsed JSON data
    except requests.exceptions.RequestException as e:
        print(f"Error fetching file: {e}")
        return None  # Return None if an error occurs

# Function to open/load local .json file(s)
def load_local_json(file_path):
    """ Load JSON data from a local file. """
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            return json.load(file)  # Read and parse JSON data
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error loading file: {e}")
        return None  # Return None if an error occurs

# Function to parse the properties from JSON
def parse_properties(json_data):
    # Extracts and returns a dictionary mapping Property IDs (PID) to their descriptions.
    return {p.get("id", "Unknown PID"): p.get("text", "No description available") for p in json_data.get("properties", [])}

# Function to parse the mitigations from JSON
def parse_mitigations(json_data):
    # Extracts and returns a dictionary mapping Mitigation IDs (MID) to their descriptions.
    return {m.get("id", "Unknown MID"): m.get("text", "No description available") for m in json_data.get("mitigations", [])}

# Function to parse the threats from JSON
def parse_threats(json_data, mitigation_dict, property_dict):
    # Extracts threats from JSON and structures them for Excel output.
    # Maps properties (PID) and mitigations (MID) to threats (TID).
    
    data_list = []
    
    # Loop through each threat entry in JSON
    for threat in json_data.get("threats", []):
        tid = threat.get("id", "Unknown TID")  # Extract Threat ID
        tid_text = threat.get("text", "No description available")  # Extract Threat Description
        properties = threat.get("properties", [])  # Get associated properties
        property_data = []

        # Extract Property IDs and Descriptions
        for prop in properties:
            pid = prop.get("id", "Unknown PID") # Extract Property ID
            pid_text = property_dict.get(pid, "No description available") # Extract Property Description
            property_data.append((pid, pid_text))

        # Extract Mitigation IDs and Descriptions
        mitigations = [(m.get("id", "Unknown MID"), mitigation_dict.get(m.get("id", "Unknown MID"), "No description available")) for m in threat.get("mitigations", [])]

        # Ensure each threat has at least one property and one mitigation
        if not property_data:
            property_data = [("None", "No associated property")]
        if not mitigations:
            mitigations = [("None", "No mitigation available")]

        # Create structured data for Excel output
        for pid, pid_text in property_data:
            for mid, mid_text in mitigations:
                data_list.append({
                    "Property ID (PID)": pid,
                    "Property Description": pid_text,  
                    "Threat ID (TID)": tid,
                    "Threat Description": tid_text,
                    "Mitigation ID (MID)": mid,
                    "Mitigation Description": mid_text
                })
    
    return data_list

def format_excel(output_file):
    """ Applies formatting to the Excel output file for better readability. """
    wb = load_workbook(output_file)
    ws = wb.active

    # Bold & enlarge headers
    header_font = Font(bold=True, size=14)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-size column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    # Merge cells for repeated values in columns
    def merge_cells(column_idx, justify="center"):
        """ Merges consecutive duplicate values in a column for better readability. """
        merge_start = None
        last_value = None

        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=column_idx).value
            ws.cell(row=row, column=column_idx).alignment = Alignment(horizontal=justify, vertical="center")

            if cell_value == last_value:
                if merge_start is None:
                    merge_start = row - 1
            else:
                if merge_start is not None:
                    ws.merge_cells(start_row=merge_start, start_column=column_idx, end_row=row - 1, end_column=column_idx)
                    merge_start = None
                last_value = cell_value

        if merge_start is not None:
            ws.merge_cells(start_row=merge_start, start_column=column_idx, end_row=ws.max_row, end_column=column_idx)

    # Apply merging to relevant columns
    merge_cells(1, "center")  # Property ID (PID)
    merge_cells(2, "left")    # Property Description
    merge_cells(3, "center")  # Threat ID (TID)
    merge_cells(4, "left")    # Threat Description
    merge_cells(6, "center")  # Mitigation ID (MID)

    wb.save(output_file)
    print(f"Formatted Excel saved as {output_file}")

def main():
    """ Main function to fetch, process, and save MITRE EMB3D threat data. """
    print("MITRE EMB3D JSON Parser")
    print("1. Use local files")
    print("2. Download from GitHub")
    choice = input("Enter your choice (1/2): ").strip()

    # Load JSON data based on user choice
    threats_json = fetch_json_from_github(THREATS_JSON_URL) if choice == "2" else load_local_json(input("Enter path to threats.json: ").strip())
    mitigations_json = fetch_json_from_github(MITIGATIONS_JSON_URL) if choice == "2" else load_local_json(input("Enter path to mitigations.json: ").strip())
    properties_json = fetch_json_from_github(PROPERTIES_JSON_URL) if choice == "2" else load_local_json(input("Enter path to properties.json: ").strip())

    parsed_data = parse_threats(threats_json, parse_mitigations(mitigations_json), parse_properties(properties_json))

    output_file = "emb3d_threats.xlsx"
    pd.DataFrame(parsed_data).to_excel(output_file, index=False, engine="openpyxl")

    if "-Format" in sys.argv:
        format_excel(output_file)

if __name__ == "__main__":
    main()
